# -*- coding: utf-8 -*-
"""Сохранение и экспорт: CSV, JSON, Excel, PDF, картинка графика.

Excel/PDF/график требуют сторонних библиотек. Их импорт обёрнут в
try/except — без библиотеки приложение работает, просто соответствующий
экспорт недоступен (об этом пользователю говорит messagebox).
"""

import os
import csv
import json
import tkinter as tk
from tkinter import filedialog, messagebox

from ..theme import colors

# openpyxl нужен только для экспорта в Excel
try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# reportlab нужен только для отчёта PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as pdf_colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ExportMixin:
    """Методы записи данных в файлы разных форматов."""

    def _ensure_data(self, title="Экспорт"):
        """Проверяет наличие данных перед экспортом."""
        if not self.project.columns:
            messagebox.showinfo(title, "Нет данных для обработки.")
            return False
        return True

    def _source_stem(self):
        """Возвращает имя открытого файла без расширения для экспортов."""
        source = self.project.meta.get("source_file")
        if source:
            return os.path.splitext(os.path.basename(source))[0]
        # файла нет — берём название проекта или запасное имя
        return self.project.meta.get("name") or "analysis"

    @staticmethod
    def _safe_filename(name):
        """Удаляет из имени файла символы, запрещённые на Windows и macOS."""
        forbidden = '<>:"/\\|?*'
        clean = "".join("_" if ch in forbidden else ch for ch in str(name))
        clean = clean.strip().strip(".")
        return clean or "analysis"

    def _export_path(self, extension, filetypes, *, suffix=""):
        """Открывает диалог сохранения с именем от текущего исходного файла."""
        initial = self._safe_filename(f"{self._source_stem()}{suffix}{extension}")
        return filedialog.asksaveasfilename(
            defaultextension=extension,
            initialfile=initial,
            filetypes=filetypes,
        )

    def _records(self):
        """Преобразует текущую таблицу в список словарей для JSON/Excel/PDF."""
        # каждая строка -> {имя_столбца: значение}
        return [dict(zip(self.project.columns, r)) for r in self.project.rows]

    def save_file(self, silent=False):
        """Сохраняет ручные изменения обратно в открытый исходный файл."""
        if not self._ensure_data("Сохранение"):
            return False
        source = self.project.meta.get("source_file")
        if not source:
            # сохранять некуда — у проекта нет исходного файла
            if not silent:
                messagebox.showinfo(
                    "Сохранение",
                    "У проекта нет исходного файла. Используйте экспорт.",
                )
            return False
        if not self.is_dirty:
            self.set_status("Нет несохранённых изменений")
            return True

        fmt = (self.project.meta.get("file_format") or "").upper()
        encoding = self.project.meta.get("encoding") or "utf-8"
        try:
            # запись зависит от исходного формата файла
            if fmt == "CSV":
                self._write_csv(source)
                encoding = "utf-8-sig"
            elif fmt == "JSON":
                with open(source, "w", encoding=encoding) as f:
                    json.dump(self._records(), f, ensure_ascii=False, indent=2)
            elif fmt == "TXT":
                self._write_txt(source, encoding)
            else:
                raise ValueError(f"Формат {fmt or 'файла'} нельзя сохранить напрямую.")
            # обновляем метаданные (размер, md5) и сбрасываем флаг изменений
            self.project._set_common_meta(source, fmt, encoding)
            self._set_dirty(False)
            self.update_meta_panel()
            self.set_status(f"Сохранено: {os.path.basename(source)}")
            return True
        except Exception as e:
            if not silent:
                messagebox.showerror("Ошибка сохранения", str(e))
            return False

    def _write_csv(self, path):
        """Записывает текущую таблицу в CSV с заголовками."""
        # utf-8-sig — с BOM, чтобы Excel правильно открывал кириллицу
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(self.project.columns)   # строка заголовков
            writer.writerows(self.project.rows)      # сами данные

    def _write_txt(self, path, encoding):
        """Записывает TXT как строки либо как табличные значения через табуляцию."""
        text_columns = ["№", "Содержимое строки"]
        if self.project.columns == text_columns:
            # файл был обычным текстом — возвращаем только содержимое строк
            lines = [
                row[1] if len(row) > 1 else ""
                for row in self.project.rows
            ]
        else:
            # таблица — склеиваем значения через табуляцию
            lines = ["\t".join(str(value) for value in row)
                     for row in self.project.rows]
        with open(path, "w", encoding=encoding) as f:
            f.write("\n".join(lines))

    def export_csv(self):
        """Экспортирует текущие данные в CSV."""
        if not self._ensure_data():
            return
        path = self._export_path(".csv", [("CSV", "*.csv")])
        if not path:
            return  # пользователь закрыл диалог
        try:
            self._write_csv(path)
            self.set_status(f"Экспортировано в CSV: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def show_export_menu(self):
        """Открывает компактное меню доступных вариантов экспорта."""
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="CSV…", command=self.export_csv)
        menu.add_command(label="JSON…", command=self.export_json)
        menu.add_command(label="Excel (xlsx)…", command=self.export_excel)
        menu.add_command(label="Отчёт PDF…", command=self.export_pdf)
        # в подписи показываем столбец, по которому будет построен график
        column = self._chart_export_column()
        chart_label = "График PNG/PDF…"
        if column:
            chart_label = f"График PNG/PDF · {column}…"
        menu.add_command(label=chart_label, command=self.export_chart_image)
        # показываем меню рядом с панелью действий
        x = self.actions.winfo_rootx() + 360
        y = self.actions.winfo_rooty() + self.actions.winfo_height()
        menu.tk_popup(x, y)

    def export_json(self):
        """Экспортирует текущие данные в JSON."""
        if not self._ensure_data():
            return
        path = self._export_path(".json", [("JSON", "*.json")])
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                # ensure_ascii=False — сохраняем кириллицу как есть
                json.dump(self._records(), f, ensure_ascii=False, indent=2)
            self.set_status(f"Экспортировано в JSON: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def export_excel(self):
        """Экспортирует текущие данные и метаданные в Excel."""
        if not self._ensure_data():
            return
        if not HAS_OPENPYXL:  # библиотека не установлена
            messagebox.showwarning(
                "Недоступно",
                "Для экспорта в Excel установите библиотеку:\n\n"
                "pip install openpyxl")
            return
        path = self._export_path(".xlsx", [("Excel", "*.xlsx")])
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Данные"
            ws.append(self.project.columns)  # заголовки в первой строке
            # заливка и белый жирный шрифт для шапки
            header_fill = PatternFill("solid", fgColor=self.pal["accent"].lstrip("#"))
            for cell in ws[1]:
                cell.font = XLFont(bold=True, color="FFFFFF")
                cell.fill = header_fill
            for r in self.project.rows:
                ws.append(r)
            # отдельный лист с метаданными проекта
            ws2 = wb.create_sheet("Метаданные")
            for k, v in self.project.meta.items():
                ws2.append([k, str(v)])
            wb.save(path)
            self.set_status(f"Экспортировано в Excel: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def export_pdf(self):
        """Создаёт PDF-отчёт анализа."""
        if not self._ensure_data():
            return
        if not HAS_REPORTLAB:
            messagebox.showwarning(
                "Недоступно",
                "Для отчёта PDF установите библиотеку:\n\n"
                "pip install reportlab")
            return
        path = self._export_path(".pdf", [("PDF", "*.pdf")], suffix="_report")
        if not path:
            return
        try:
            self._build_pdf(path)
            self.set_status(f"Отчёт сохранён: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _register_pdf_font(self):
        """Подбирает шрифт с кириллицей. Возвращает имя шрифта."""
        # перебираем типичные пути к шрифтам на Windows/macOS/Linux
        candidates = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/Library/Fonts/Arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for fp in candidates:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("Custom", fp))
                    return "Custom"
                except Exception:
                    continue
        return "Helvetica"  # запасной встроенный шрифт (без кириллицы)

    def _build_pdf(self, path):
        """Формирует структуру PDF-отчёта."""
        font = self._register_pdf_font()
        pal = colors.palette(self.theme_name)
        # документ A4 с полями 1.5 см со всех сторон
        doc = SimpleDocTemplate(path, pagesize=A4,
                                leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                                topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles_sheet = getSampleStyleSheet()
        # подменяем шрифты у стандартных стилей на наш (с кириллицей)
        styles_sheet["Title"].fontName = font
        styles_sheet["Normal"].fontName = font
        styles_sheet["Heading2"].fontName = font

        # заголовок + блок метаданных
        elements = [Paragraph("Отчёт анализа файла", styles_sheet["Title"]),
                    Spacer(1, 0.4 * cm),
                    Paragraph("Метаданные проекта", styles_sheet["Heading2"])]

        meta_rows = [[k, str(v)] for k, v in self.project.meta.items()]
        mt = Table(meta_rows, colWidths=[4 * cm, 13 * cm])
        mt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, pdf_colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), pdf_colors.HexColor(pal["heading"])),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements += [mt, Spacer(1, 0.5 * cm),
                     Paragraph("Данные (первые 100 строк)", styles_sheet["Heading2"])]

        # в отчёт кладём не более 8 столбцов и 100 строк (чтобы влезло)
        max_cols = 8
        cols = self.project.columns[:max_cols]
        table_data = [cols]
        for r in self.project.rows[:100]:
            table_data.append([self._truncate(r[i] if i < len(r) else "", 28)
                               for i in range(len(cols))])
        dt = Table(table_data, repeatRows=1)  # repeatRows=1 — шапка на каждой странице
        dt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, pdf_colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor(pal["accent"])),
            ("TEXTCOLOR", (0, 0), (-1, 0), pdf_colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [pdf_colors.white, pdf_colors.HexColor(pal["stripe"])]),
        ]))
        elements.append(dt)
        # если столбцов было больше — добавляем сноску
        if len(self.project.columns) > max_cols:
            elements += [Spacer(1, 0.3 * cm),
                         Paragraph(f"* Показаны первые {max_cols} столбцов из "
                                   f"{len(self.project.columns)}.",
                                   styles_sheet["Normal"])]
        doc.build(elements)  # собираем PDF из подготовленных элементов

    @staticmethod
    def _truncate(v, n):
        """Обрезает длинную строку до заданной длины."""
        s = str(v)
        return s if len(s) <= n else s[:n - 1] + "…"

    def export_chart_image(self):
        """Экспортирует график по последнему столбцу окна статистики."""
        if not self.project.columns:
            messagebox.showinfo("Нет данных", "Сначала загрузите данные.")
            return
        col = self._chart_export_column()
        if col is None:
            return
        self._save_chart(self._model_for(col), col)

    def _save_chart(self, model, column):
        """Сохраняет модель графика в PNG или PDF."""
        from ..ui import chart  # импорт здесь, чтобы не тянуть PIL без надобности
        if not chart.HAS_PIL:
            messagebox.showwarning(
                "Недоступно",
                "Для экспорта графиков установите библиотеку:\n\n"
                "pip install Pillow")
            return
        suffix = f"_chart_{self._safe_filename(column)}"
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            initialfile=self._safe_filename(f"{self._source_stem()}{suffix}.png"),
            filetypes=[("PNG", "*.png"), ("PDF", "*.pdf")])
        if not path:
            return
        try:
            chart.render_image(model, colors.palette(self.theme_name), path)
            self.set_status(
                f"График по столбцу «{column}» сохранён: {os.path.basename(path)}"
            )
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
